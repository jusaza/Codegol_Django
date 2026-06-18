import csv
from django.contrib import messages

from django.shortcuts import render, get_object_or_404, redirect
from .models import Matricula, HistorialCategoria
from usuario.models import Usuario, DetallesUsuarioRol
from categoria.models import Categoria
from datetime import date
from .models import Matricula, HistorialCategoria
from django.db.models import Q
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from django.http import HttpResponse
from posicion.models import Posicion

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)
from openpyxl.utils import get_column_letter
from django.db.models import Q

def cargar_matriculas_csv(request):

    if request.method == "POST":

        archivo = request.FILES.get('archivo')

        if not archivo:
            messages.error(
                request,
                "Debes seleccionar un archivo CSV."
            )
            return redirect(
                'carga_masiva_matricula'
            )

        decoded_file = (
            archivo.read()
            .decode('utf-8-sig')
            .splitlines()
        )

        reader = csv.DictReader(decoded_file)

        columnas_requeridas = [
            'num_identificacion',
            'fecha_inicio',
            'fecha_fin',
            'nivel',
            'observaciones',
            'posicion',
            'categoria'
        ]

        columnas_faltantes = [
            c for c in columnas_requeridas
            if c not in reader.fieldnames
        ]

        if columnas_faltantes:

            messages.error(
                request,
                f"❌ Faltan columnas: {', '.join(columnas_faltantes)}"
            )

            return redirect(
                'carga_masiva_matricula'
            )

        matriculas_data = []

        errores = []

        for fila, row in enumerate(reader, start=2):

            documento = row["num_identificacion"]

            # ===========================
            # JUGADOR
            # ===========================

            try:

                jugador = Usuario.objects.get(
                    num_identificacion=documento
                )

            except Usuario.DoesNotExist:

                errores.append(
                    f"Fila {fila}: Usuario no existe ({documento})"
                )

                continue

            # ===========================
            # FECHAS
            # ===========================

            try:

                fecha_inicio = datetime.strptime(
                    row["fecha_inicio"],
                    "%Y-%m-%d"
                ).date()

                fecha_fin = datetime.strptime(
                    row["fecha_fin"],
                    "%Y-%m-%d"
                ).date()

            except ValueError:

                errores.append(
                    f"Fila {fila}: Formato de fecha inválido ({documento})"
                )

                continue

            if fecha_inicio > fecha_fin:

                errores.append(
                    f"Fila {fila}: La fecha inicio es mayor que la fecha fin ({documento})"
                )

                continue

            # ===========================
            # VALIDAR SOLAPAMIENTO
            # ===========================

            existe = Matricula.objects.filter(
                id_jugador=jugador,
                estado=True,
                fecha_inicio__lte=fecha_fin,
                fecha_fin__gte=fecha_inicio
            ).exists()

            if existe:

                errores.append(
                    f"Fila {fila}: Ya existe una matrícula que se cruza con esas fechas ({documento})"
                )

                continue

            # ===========================
            # POSICIÓN
            # ===========================

            try:

                posicion = Posicion.objects.get(
                    nombre=row["posicion"]
                )

            except Posicion.DoesNotExist:

                errores.append(
                    f"Fila {fila}: Posición inválida ({row['posicion']})"
                )

                continue

            # ===========================
            # CREAR OBJETO
            # ===========================

            matricula = Matricula(

                fecha_inicio=fecha_inicio,

                fecha_fin=fecha_fin,

                nivel=row["nivel"],

                observaciones=row["observaciones"],

                id_jugador=jugador,

                posicion=posicion,

                estado=True

            )

            matriculas_data.append({

                "matricula": matricula,

                "categoria": row["categoria"]

            })

        # ===========================
        # GUARDAR MATRÍCULAS
        # ===========================

        objs = [
            x["matricula"]
            for x in matriculas_data
        ]

        Matricula.objects.bulk_create(objs)

        # ===========================
        # HISTORIAL DE CATEGORÍA
        # ===========================

        for item in matriculas_data:

            matricula = item["matricula"]

            categoria_nombre = item["categoria"]

            try:

                categoria = Categoria.objects.get(
                    nombre_categoria=categoria_nombre
                )

                HistorialCategoria.objects.create(

                    id_matricula=matricula,

                    id_categoria=categoria,

                    fecha_registro=matricula.fecha_inicio,

                    estado=True

                )

            except Categoria.DoesNotExist:

                errores.append(
                    f"No existe la categoría: {categoria_nombre}"
                )

        # ===========================
        # MENSAJES
        # ===========================

        if matriculas_data:

            messages.success(
                request,
                f"✅ Matrículas creadas: {len(matriculas_data)}"
            )

        if errores:

            messages.error(
                request,
                f"❌ Registros omitidos ({len(errores)}):\n"
                + "\n".join(errores[:10])
            )

        return redirect(
            'carga_masiva_matricula'
        )

    return render(
        request,
        "matricula/cargar.html"
    )

def lista_matricula(request):
    query = request.GET.get('q')
    usuario_id = request.session.get("usuario_id")
    es_jugador = DetallesUsuarioRol.objects.filter(
        id_usuario_id=usuario_id,
        id_rol__rol_usuario__iexact="Jugador"
    ).exists()
    if es_jugador:
        matriculas = Matricula.objects.filter(
            estado=True,
            id_jugador__id_usuario=usuario_id  
        )
    else:
        matriculas = Matricula.objects.filter(estado=True)
    if query:
        matriculas = matriculas.filter(
            Q(id_jugador__nombre_completo__icontains=query) |
            Q(id__icontains=query)
        )
    for m in matriculas:
        ultima = HistorialCategoria.objects.filter(
            id_matricula=m,
            estado=True
        ).order_by('-fecha_registro', '-id_historial').first()
        m.categoria_actual = (
            ultima.id_categoria.nombre_categoria
            if ultima else "Sin categoría"
        )
    return render(request, 'matricula/lista.html', {
        'matriculas': matriculas,
        'query': query
    })

# CREAR
def crear_matricula(request):

    hoy = date.today()

    usuarios = Usuario.objects.filter(
        roles__rol_usuario="Jugador"
    ).exclude(
        matriculas__estado=True,
        matriculas__fecha_fin__gte=hoy
    ).distinct()

    posiciones = Posicion.objects.all()

    categorias = Categoria.objects.filter(
        estado=True
    )

    if request.method == 'POST':

        # ==========================
        # DATOS DEL FORMULARIO
        # ==========================

        fecha_inicio = datetime.strptime(
            request.POST.get('fecha_inicio'),
            "%Y-%m-%d"
        ).date()

        fecha_fin = datetime.strptime(
            request.POST.get('fecha_fin'),
            "%Y-%m-%d"
        ).date()

        fecha_matricula = request.POST.get(
            'fecha_matricula'
        )

        nivel = request.POST.get(
            'nivel'
        )

        observaciones = request.POST.get(
            'observaciones'
        )

        id_jugador = request.POST.get(
            'id_jugador'
        )

        posicion_id = request.POST.get(
            'posicion'
        )

        categoria_id = request.POST.get(
            'categoria'
        )

        # ==========================
        # VALIDACIONES
        # ==========================

        if not id_jugador:

            return render(
                request,
                'matricula/form.html',
                {
                    'usuarios': usuarios,
                    'posiciones': posiciones,
                    'categorias': categorias,
                    'error': 'Debe seleccionar un jugador'
                }
            )

        if fecha_inicio > fecha_fin:

            return render(
                request,
                'matricula/form.html',
                {
                    'usuarios': usuarios,
                    'posiciones': posiciones,
                    'categorias': categorias,
                    'error': 'La fecha de inicio no puede ser mayor que la fecha final.'
                }
            )

        jugador = Usuario.objects.get(
            id_usuario=id_jugador
        )

        # ==========================
        # VALIDAR SOLAPAMIENTO
        # ==========================

        existe_solapamiento = Matricula.objects.filter(
            id_jugador=jugador,
            estado=True
        ).filter(
            fecha_inicio__lte=fecha_fin,
            fecha_fin__gte=fecha_inicio
        ).exists()

        if existe_solapamiento:

            return render(
                request,
                'matricula/form.html',
                {
                    'usuarios': usuarios,
                    'posiciones': posiciones,
                    'categorias': categorias,
                    'error': (
                        'El jugador ya tiene una matrícula '
                        'vigente durante ese período. '
                        'Debe esperar a que finalice la '
                        'matrícula actual para registrar una nueva.'
                    )
                }
            )

        # ==========================
        # CREAR MATRÍCULA
        # ==========================

        matricula = Matricula.objects.create(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            fecha_matricula=fecha_matricula,
            nivel=nivel,
            observaciones=observaciones,
            estado=True,
            id_jugador=jugador,
            posicion=Posicion.objects.get(
                id_posicion=posicion_id
            ) if posicion_id else None
        )

        # ==========================
        # CATEGORÍA
        # ==========================

        if categoria_id:

            HistorialCategoria.objects.create(
                id_matricula=matricula,
                id_categoria_id=categoria_id,
                fecha_registro=date.today(),
                estado=True
            )

        return redirect(
            'lista_matricula'
        )

    return render(
        request,
        'matricula/form.html',
        {
            'usuarios': usuarios,
            'posiciones': posiciones,
            'categorias': categorias
        }
    )


# EDITAR
def editar_matricula(request, id):

    matricula = get_object_or_404(
        Matricula,
        id=id
    )

    posiciones = Posicion.objects.all()

    categorias = Categoria.objects.filter(
        estado=True
    )

    ultima_categoria = (
        HistorialCategoria.objects
        .filter(
            id_matricula=matricula,
            estado=True
        )
        .order_by(
            '-fecha_registro',
            '-id_historial'
        )
        .first()
    )

    if request.method == 'POST':

        # ==========================
        # DATOS MATRÍCULA
        # ==========================

        matricula.fecha_inicio = request.POST.get(
            'fecha_inicio'
        )

        matricula.fecha_fin = request.POST.get(
            'fecha_fin'
        )

        matricula.fecha_matricula = request.POST.get(
            'fecha_matricula'
        )

        matricula.nivel = request.POST.get(
            'nivel'
        )

        matricula.observaciones = request.POST.get(
            'observaciones'
        )

        # ==========================
        # POSICIÓN
        # ==========================

        posicion_id = request.POST.get(
            'posicion'
        )

        if posicion_id:

            matricula.posicion = Posicion.objects.get(
                id_posicion=posicion_id
            )

        matricula.save()

        # ==========================
        # CATEGORÍA
        # ==========================

        categoria_nueva = request.POST.get(
            'categoria'
        )

        observacion_categoria = request.POST.get(
            'observacion_categoria'
        )

        categoria_actual = None

        if ultima_categoria:

            categoria_actual = str(
                ultima_categoria.id_categoria.id_categoria
            )

        # ==========================
        # SI CAMBIÓ LA CATEGORÍA
        # ==========================

        if (
            categoria_nueva and
            categoria_actual != str(categoria_nueva)
        ):

            # Desactivar categorías activas anteriores
            HistorialCategoria.objects.filter(
                id_matricula=matricula,
                estado=True
            ).update(
                estado=False
            )

            # Crear nueva categoría activa
            HistorialCategoria.objects.create(
                id_matricula=matricula,
                id_categoria_id=categoria_nueva,
                fecha_registro=date.today(),
                observacion=observacion_categoria,
                estado=True
            )

        return redirect(
            'lista_matricula'
        )

    return render(
        request,
        'matricula/form.html',
        {
            'matricula': matricula,
            'posiciones': posiciones,
            'categorias': categorias,
            'ultima_categoria': ultima_categoria
        }
    )


def eliminar_matricula(request, id):
    matricula = get_object_or_404(Matricula, id=id)
    matricula.estado = False
    matricula.save()

    return redirect('lista_matricula')



def asignar_categoria(request, id):

    matricula = get_object_or_404(
        Matricula,
        id=id
    )

    categorias = Categoria.objects.filter(
        estado=True
    )

    if request.method == 'POST':

        id_categoria = request.POST.get(
            'categoria'
        )

        if id_categoria:

            # Desactivar categoría activa actual
            HistorialCategoria.objects.filter(
                id_matricula=matricula,
                estado=True
            ).update(
                estado=False
            )

            # Crear nueva categoría activa
            HistorialCategoria.objects.create(
                id_matricula=matricula,
                id_categoria_id=id_categoria,
                fecha_registro=date.today(),
                estado=True
            )

        return redirect(
            'lista_matricula'
        )

    return render(
        request,
        'matricula/asignar_categoria.html',
        {
            'matricula': matricula,
            'categorias': categorias
        }
    )

def ver_historial_categoria(request, id):
    matricula = get_object_or_404(Matricula, id=id)

    historial = HistorialCategoria.objects.filter(
        id_matricula=matricula
    ).order_by('-fecha_registro')

    return render(request, 'matricula/historial_categoria.html', {
        'matricula': matricula,
        'historial': historial
    })

def generar_certificado(request, id):

    matricula = Matricula.objects.get(id=id)
    jugador = matricula.id_jugador
    historial = HistorialCategoria.objects.filter(
        id_matricula=matricula
    ).order_by('fecha_registro')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="certificado_matricula.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()

    contenido = []

    # TITULO
    contenido.append(Paragraph("CERTIFICADO DE MATRÍCULA DEPORTIVA", styles['Title']))
    contenido.append(Spacer(1, 30))

    # TEXTO PRINCIPAL
    texto = f"""
    La presente certifica que el jugador <b>{jugador.nombre_completo}</b>,
    identificado con <b>{jugador.get_tipo_documento_display()}</b> número
    <b>{jugador.num_identificacion}</b>, se encuentra vinculado a la institución
    deportiva, habiendo formalizado su matrícula el día
    <b>{matricula.fecha_inicio}</b>.
    
    Durante su permanencia en la escuela, el jugador ha participado activamente
    en los procesos formativos correspondientes, demostrando compromiso,
    disciplina y desarrollo deportivo.
    """

    contenido.append(Paragraph(texto, styles['Normal']))
    contenido.append(Spacer(1, 25))

    # HISTORIAL
    contenido.append(Paragraph("Historial de Categorías", styles['Heading2']))
    contenido.append(Spacer(1, 15))

    if historial.exists():
        for h in historial:
            contenido.append(
                Paragraph(
                    f"• {h.id_categoria.nombre_categoria} "
                    f"({h.fecha_registro})",
                    styles['Normal']
                )
            )
    else:
        contenido.append(Paragraph("• No registra categorías asignadas", styles['Normal']))

    contenido.append(Spacer(1, 40))

    # CIERRE
    cierre = """
    Este certificado se expide a solicitud del interesado para los fines que estime convenientes.
    """

    contenido.append(Paragraph(cierre, styles['Normal']))
    contenido.append(Spacer(1, 60))

    # FIRMA
    contenido.append(Paragraph("__________________________________", styles['Normal']))
    contenido.append(Paragraph("Dirección Deportiva", styles['Normal']))
    contenido.append(Paragraph("Escuela de Formación", styles['Normal']))

    doc.build(contenido)

    return response

def modal_filtro_excel(request):

    categorias = Categoria.objects.filter(estado=True)
    posiciones = Posicion.objects.all()

    return render(request, 'matricula/modal_filtro_excel.html', {
        'categorias': categorias,
        'posiciones': posiciones
    })

def exportar_matriculas_excel(request):

    matriculas = Matricula.objects.filter(
        estado=True
    )

    categoria = request.GET.get('categoria')
    nivel = request.GET.get('nivel')
    posicion = request.GET.get('posicion')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    # =========================
    # FILTROS
    # =========================

    if nivel:
        matriculas = matriculas.filter(
            nivel=nivel
        )

    if posicion:
        matriculas = matriculas.filter(
            posicion_id=posicion
        )

    if fecha_inicio:
        matriculas = matriculas.filter(
            fecha_inicio__gte=fecha_inicio
        )

    if fecha_fin:
        matriculas = matriculas.filter(
            fecha_fin__lte=fecha_fin
        )

    if categoria:

        matriculas = matriculas.filter(
            historialcategoria__id_categoria_id=categoria,
            historialcategoria__estado=True
        ).distinct()

    total = matriculas.count()

    # =========================
    # EXCEL
    # =========================

    wb = Workbook()

    ws = wb.active
    ws.title = "Matrículas"

    # =========================
    # ESTILOS
    # =========================

    titulo_fill = PatternFill(
        "solid",
        fgColor="0D2B4D"
    )

    header_fill = PatternFill(
        "solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    titulo_font = Font(
        bold=True,
        color="FFFFFF",
        size=16
    )

    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # =========================
    # TITULO
    # =========================

    ws.merge_cells("A1:I1")

    ws["A1"] = "REPORTE DE MATRÍCULAS"

    ws["A1"].fill = titulo_fill
    ws["A1"].font = titulo_font

    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    # =========================
    # RESUMEN
    # =========================

    ws["A3"] = "Fecha de generación"
    ws["B3"] = datetime.now().strftime(
        "%d/%m/%Y %H:%M"
    )

    ws["A4"] = "Total jugadores"
    ws["B4"] = total

    ws["A5"] = "Nivel"
    ws["B5"] = nivel or "Todos"

    ws["A6"] = "Posición"
    ws["B6"] = posicion or "Todas"

    ws["A7"] = "Categoría"
    ws["B7"] = categoria or "Todas"

    ws["A8"] = "Fecha Inicio"
    ws["B8"] = fecha_inicio or "Todas"

    ws["A9"] = "Fecha Fin"
    ws["B9"] = fecha_fin or "Todas"

    for fila in range(3, 10):

        ws[f"A{fila}"].font = Font(
            bold=True
        )

        ws[f"A{fila}"].fill = PatternFill(
            "solid",
            fgColor="D9EAD3"
        )

    # =========================
    # ENCABEZADOS
    # =========================

    fila_header = 11

    headers = [
        "ID",
        "Jugador",
        "Tipo Documento",
        "Número Documento",
        "Nivel",
        "Fecha Inicio",
        "Fecha Fin",
        "Fecha Matrícula",
        "Observaciones"
    ]

    for col_num, texto in enumerate(
        headers,
        start=1
    ):

        cell = ws.cell(
            row=fila_header,
            column=col_num
        )

        cell.value = texto
        cell.fill = header_fill
        cell.font = header_font
        cell.border = borde

        cell.alignment = Alignment(
            horizontal="center"
        )

    # =========================
    # DATOS
    # =========================

    fila = 12

    for m in matriculas:

        jugador = m.id_jugador

        datos = [
            m.id,
            jugador.nombre_completo,
            jugador.get_tipo_documento_display(),
            jugador.num_identificacion,
            m.nivel,
            str(m.fecha_inicio),
            str(m.fecha_fin),
            str(m.fecha_matricula),
            m.observaciones or ""
        ]

        for col_num, valor in enumerate(
            datos,
            start=1
        ):

            cell = ws.cell(
                row=fila,
                column=col_num
            )

            cell.value = valor
            cell.border = borde

        fila += 1

    # =========================
    # AJUSTAR COLUMNAS
    # =========================

    for columna in ws.columns:

        try:

            longitud = 0

            letra = get_column_letter(
                columna[0].column
            )

            for celda in columna:

                if celda.value:

                    longitud = max(
                        longitud,
                        len(str(celda.value))
                    )

            ws.column_dimensions[
                letra
            ].width = longitud + 5

        except:
            pass

    # =========================
    # RESPUESTA
    # =========================

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; '
        'filename="reporte_matriculas.xlsx"'
    )

    wb.save(response)

    return response